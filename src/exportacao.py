"""
Módulo de exportação.
Gera a planilha de resultado com aba CRÍTICA + uma aba por conferência.
"""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Cores por status
CORES = {
    "✅ OK":                    "C6EFCE",  # verde claro
    "🚨 NÃO LANÇADO":           "FFC7CE",  # vermelho claro
    "🚨 LANÇAMENTO INDEVIDO":   "FFC7CE",  # vermelho claro
    "🟡 DE FÉRIAS":             "FFEB9C",  # amarelo claro
    "🟡 POSTERGADO (FÉRIAS)":   "FFEB9C",  # amarelo claro
}

COR_HEADER = "1F4E79"  # azul escuro
COR_HEADER_FONT = "FFFFFF"


def _aplicar_estilo(ws, df: pd.DataFrame):
    """Aplica estilo visual à planilha."""
    # Header
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True, color=COR_HEADER_FONT, size=11)
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        status = ""
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="center")
            col_name = df.columns[col_idx - 1]
            if col_name == "STATUS":
                status = str(value)
        # Colorir linha inteira com base no status
        cor = CORES.get(status, "FFFFFF")
        fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Ajustar largura das colunas
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Congelar linha do header
    ws.freeze_panes = "A2"


def exportar_resultado(resultados: dict, caminho_saida: str, mes_referencia: str):
    """
    Exporta todos os resultados para um arquivo Excel.

    Args:
        resultados: dict com chave = nome da aba, valor = DataFrame
        caminho_saida: caminho do arquivo de saída (.xlsx)
        mes_referencia: string 'MM/AAAA' do mês de referência
    """
    # Montar aba CRÍTICA (apenas os não-OK)
    criticos = []
    for nome_aba, df in resultados.items():
        if df.empty:
            continue
        df_critico = df[~df["STATUS"].str.startswith("✅")].copy()
        if not df_critico.empty:
            df_critico.insert(0, "CONFERENCIA", nome_aba)
            criticos.append(df_critico)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        # Aba CRÍTICA primeiro
        if criticos:
            df_critica = pd.concat(criticos, ignore_index=True)
        else:
            df_critica = pd.DataFrame(columns=["CONFERENCIA", "STATUS"])

        df_critica.to_excel(writer, sheet_name="CRÍTICA", index=False)
        _aplicar_estilo(writer.sheets["CRÍTICA"], df_critica)

        # Demais abas
        for nome_aba, df in resultados.items():
            if df.empty:
                continue
            nome_aba_limpo = nome_aba[:31]  # Excel limita a 31 chars
            df.to_excel(writer, sheet_name=nome_aba_limpo, index=False)
            _aplicar_estilo(writer.sheets[nome_aba_limpo], df)

    print(f"✅ Relatório exportado: {caminho_saida}")
